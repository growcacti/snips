import tkinter as tk
from openpyxl import load_workbook

def insert_data():
    # Load the workbook and select the specific worksheet
    workbook = load_workbook('example.xlsx')
    sheet_name = sheet_name_entry.get()
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Insert data into the specified cell
        cell = cell_entry.get()
        data = data_entry.get()
        sheet[cell] = data

        # Save the workbook
        workbook.save('example.xlsx')
        status_label.config(text="Data inserted successfully.")
    else:
        status_label.config(text="Sheet not found.")

# Setting up the Tkinter window
root = tk.Tk()
root.title("Excel Data Inserter")

tk.Label(root, text="Sheet Name:").grid(row=0, column=0)
sheet_name_entry = tk.Entry(root)
sheet_name_entry.grid(row=0, column=1)

tk.Label(root, text="Cell:").grid(row=1, column=0)
cell_entry = tk.Entry(root)
cell_entry.grid(row=1, column=1)

tk.Label(root, text="Data:").grid(row=2, column=0)
data_entry = tk.Entry(root)
data_entry.grid(row=2, column=1)

insert_button = tk.Button(root, text="Insert Data", command=insert_data)
insert_button.grid(row=3, column=0, columnspan=2)

status_label = tk.Label(root, text="")
status_label.grid(row=4, column=0, columnspan=2)

# Adjust spacing
root.grid_columnconfigure(0, pad=10)
root.grid_columnconfigure(1, pad=10)
root.grid_rowconfigure(0, pad=5)
root.grid_rowconfigure(1, pad=5)
root.grid_rowconfigure(2, pad=5)
root.grid_rowconfigure(3, pad=5)

root.mainloop()
